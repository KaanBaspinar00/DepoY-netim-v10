import pandas as pd
"""
# Define the product list
products = [
    "50 GR 165 CM ELYAF", "60 GR 65 CM ELYAF", "60 GR 140 CM ELYAF", "60 GR 150 CM ELYAF", "60 GR 160 CM ELYAF",
    "60 GR 210 CM ELYAF", "60 GR 240 CM ELYAF", "80 GR 70 CM ELYAF", "80 GR 75 CM ELYAF", "80 GR 85 CM ELYAF",
    "100 GR 75 CM ELYAF", "100 GR 70 CM ELYAF", "100 GR 210 CM ELYAF", "120 GR 85 CM ELYAF",
    "120 GR 160 CM ELYAF",
    "210 GÜLLÜ JAGAR", "210 YILANLI JAGAR", "210 DİAGONAL JAGAR", "210 NOKTALI JAGAR", "210 BAKLAVA JAGAR",
    "210 ÜÇ ÇİZGİLİ JAGAR", "210 EKRU 3 ÇİZGİLİ BEKART", "210 GRİ 3 ÇİZGİLİ BEKART", "210 EKRU DÜZ BEKART",
    "80 GR 240 CM MİKRO", "80 GR 80 CM MİKRO", "80 GR 90 CM MİKRO", "80 GR 75 CM MİKRO", "100 GR 300 CM MİKRO",
    "100 GR 90 CM MİKRO", "100 GR 80 CM MİKRO", "100 GR 75 CM MİKRO", "220 CM ASTAR", "46 CM BASKISIZ KOLİ",
    "30 CM BASKISIZ KOLİ", "35 CM BASKISIZ KOLİ", "42 CM BASKILI KOLİ", "50 CM BASKILI KOLİ", "GÖMLEK KOLİSİ",
    "ÇARŞAF KOLİSİ", "DANTEL SARIM KOLİSİ", "220 CM ASTAR", "250 CM BEYAZ İP", "160 CM ASTAR",
    "240 CM BEYAZ İP",
    "80 CM BEYAZ İP", "160 CM BEYAZ İP", "90 CM BEYAZ İP", "5,5 CM ASTAR", "75 CM ASTAR",
    "160 CM POLY PUANTİYE",
    "80 CM POLY PUANTİYE", "4,5 CM POLY PUANTİYE", "60 GR 70 CM TELA", "60 GR 80 CM TELA", "60 GR 160 CM TELA",
    "80 GR 160 CM TELA", "80 GR 80 CM TELA", "40 GR 160 CM TELA", "40 GR 120 CM TELA", "40 GR 90 CM TELA",
    "40 GR 67 CM TELA", "40 GR 80 CM TELA", "60 GR 65 CM TELA", "15 GR 75 CM TELA", "15 GR 80 CM TELA",
    "15 GR 210 CM TELA", "75 CM ASTAR", "75 CM ASTAR", "80 CM ASTAR", "90 CM ASTAR", "165 CM ASTAR",
    "210 CM ASTAR", "60 CM ASTAR", "65 CM ASTAR", "330 CM ASTAR", "280 CM ASTAR", "300 CM ASTAR",
    "280 CM ŞEKER KASAR", "90 CM ŞEKER KASAR", "80 CM ŞEKER KASAR", "5,5 CM ŞEKER KASAR"
]

# Create a DataFrame with columns 'Varlık adı' and 'minimum stock'
df = pd.DataFrame({
    'Varlık adı': products,
    'minimum stock': 1
})

# Save the DataFrame to an Excel file
file_path = "Veriler/Malzeme Uyarı.xlsx"
df.to_excel(file_path, index=False)"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.units import pixels_to_points

# Constants for tag dimensions and page setup
TAG_WIDTH_MM = 105  # in millimeters
TAG_HEIGHT_MM = 57  # in millimeters
TAG_SPACING_MM = 4.5  # in millimeters
PAGE_WIDTH_MM = 210  # A4 width in millimeters
PAGE_HEIGHT_MM = 297  # A4 height in millimeters

# Conversion factor: 1 mm ≈ 3.78 pixels (approximation for Excel)
MM_TO_PIXELS = 3.78

# Calculate tag and spacing dimensions in pixels
tag_width_pixels = int(TAG_WIDTH_MM * MM_TO_PIXELS)
tag_height_pixels = int(TAG_HEIGHT_MM * MM_TO_PIXELS)
spacing_pixels = int(TAG_SPACING_MM * MM_TO_PIXELS)
page_width_pixels = int(PAGE_WIDTH_MM * MM_TO_PIXELS)
page_height_pixels = int(PAGE_HEIGHT_MM * MM_TO_PIXELS)

# Create a workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Tag Template"

# Remove default headers and footers
ws.oddHeader.left.text = ""
ws.oddFooter.left.text = ""

# Styling for borders and alignment
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                     bottom=Side(style='thin'))
alignment = Alignment(horizontal='center', vertical='center')

# Calculate number of tags per row and column
tags_per_row = page_width_pixels // (tag_width_pixels + spacing_pixels)
tags_per_col = page_height_pixels // (tag_height_pixels + spacing_pixels)

# Populate the worksheet with tags
row_offset = 1  # Starting row
for i in range(tags_per_col):
    for j in range(tags_per_row):
        # Calculate cell positions for the tag
        start_row = row_offset + i * (tag_height_pixels + spacing_pixels)
        start_col = 1 + j * (tag_width_pixels + spacing_pixels)

        # Merge cells for the tag
        end_row = start_row + tag_height_pixels
        end_col = start_col + tag_width_pixels
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

        # Style the tag area
        cell = ws.cell(row=start_row, column=start_col)
        cell.border = thin_border
        cell.alignment = alignment
        cell.value = f"Tag ({i + 1}, {j + 1})"

# Save the workbook
output_file = "tag_template.xlsx"
wb.save(output_file)




