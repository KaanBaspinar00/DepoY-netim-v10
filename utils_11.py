import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
import uuid
import streamlit as st
import qrcode
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
import json
from io import BytesIO
from datetime import datetime

# ---------------------- File Paths ----------------------

# Define all necessary file paths
main_excel_file = "Veriler/depo_veri12.xlsx"
qr_images_file = "Veriler/qr_images12.xlsx"
recent_qr_codes_file = "Veriler/recent_qr_codes12.xlsx"
qr_codes_folder = "Veriler/qr_codes12"
asset_movements_excel = "Veriler/asset_movements1.xlsx"
qr_codes_output_excel = "Veriler/qr_codes_output.xlsx"
malzeme_uyari_file = "Veriler/Malzeme Uyarı.xlsx"
tasks_excel_file = "Veriler/tasks.xlsx"

# Ensure the QR codes folder exists
os.makedirs(qr_codes_folder, exist_ok=True)

# ---------------------- Excel File Initialization ----------------------

def create_excel_file_if_missing(file_path, headers):
    """Create an Excel file with specified headers if it doesn't exist."""
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(file_path)
        print(f"Created missing Excel file: {file_path}")


# Initialize Excel files with correct headers
create_excel_file_if_missing(main_excel_file,
                             ["id", "qr_kodu", "gönderen", "alıcı", "varlık_adı", "miktar", "unit", "adet", "kacıncı",
                              "zaman", "quantity"])
create_excel_file_if_missing(recent_qr_codes_file, ["id", "QR-codes-text", "image_path"])  # Updated key name
create_excel_file_if_missing(asset_movements_excel,
                             ["movement_id", "Zaman", "Varlık ID", "Varlık Adı", "Aksiyon", "Miktar", "Firma",
                              "Çalışan", "Notlar"])
create_excel_file_if_missing(tasks_excel_file,
                             ["task_id", "title", "description", "assigned_to", "created_by", "urgency", "status",
                              "progress", "created_at", "updated_at"])


# ---------------------- Helper Functions ----------------------

def authenticate_user(username, password):
    """Authenticate user based on roles.txt file."""
    try:
        with open("Veriler/roles.txt", "r", encoding='utf-8') as file:
            user_data = json.load(file)  # Load data as JSON
            users = user_data.get("users", {})
            user_info = users.get(username)
            if user_info and user_info.get("password") == password:
                return user_info.get("role")
    except json.JSONDecodeError:
        st.error("Error reading Veriler/roles.txt: Invalid JSON format.")
    except FileNotFoundError:
        st.error("Veriler/roles.txt file not found.")
    except Exception as e:
        st.error(f"An unexpected error occurred during authentication: {e}")
    return None


def add_asset_to_excel(id_, qr_kodu, varlık_adı, gönderen, alıcı, miktar, unit, adet, kacıncı, zaman):
    """Append or update asset information in the main Excel file."""
    try:
        if os.path.exists(main_excel_file):
            assets_df = pd.read_excel(main_excel_file)
        else:
            assets_df = pd.DataFrame(
                columns=["id", "qr_kodu", "gönderen", "alıcı", "varlık_adı", "miktar", "unit", "adet", "kacıncı",
                         "zaman", "quantity"])

        # Check if asset exists
        if id_ in assets_df['id'].values:
            # Update existing asset's 'miktar' and 'quantity'
            assets_df.loc[assets_df['id'] == id_, 'miktar'] = miktar
            assets_df.loc[assets_df['id'] == id_, 'quantity'] = miktar
        else:
            # Create a new asset entry
            new_asset = {
                'id': id_,
                'qr_kodu': qr_kodu,
                'gönderen': gönderen,
                'alıcı': alıcı,
                'varlık_adı': varlık_adı,
                'miktar': miktar,
                'unit': unit,
                'adet': adet,
                'kacıncı': kacıncı,
                'zaman': zaman,
                'quantity': miktar  # Initialize quantity
            }
            # Replace append() with pd.concat()
            assets_df = pd.concat([assets_df, pd.DataFrame([new_asset])], ignore_index=True)

        assets_df.to_excel(main_excel_file, index=False)

        # Update 'Malzeme Uyarı.xlsx' after saving the main Excel file
        update_malzeme_uyari()
    except Exception as e:
        st.error(f"Error adding asset to Excel: {e}")


def generate_qr_codes(assetname, unit, miktar, adet, gönderen, alıcı):
    """Generate QR codes and save them into the Excel template."""
    try:
        zaman = datetime.now().strftime("%Y-%m-%d-%H-%M-%S-%f")
        qr_codes_info = []

        for kacıncı in range(1, adet + 1):
            uuid_part = uuid.uuid4().hex
            unique_id = f"{assetname}-{gönderen}-{alıcı}-{miktar}{unit}-{adet}-{kacıncı}-{zaman}-{uuid_part}"
            qr_code_text = unique_id

            # Generate and save the QR code image
            qr = qrcode.make(qr_code_text)
            qr_image_path = os.path.join(qr_codes_folder, f"{unique_id}.png")
            qr.save(qr_image_path)

            # Collect QR code info with consistent key names
            qr_codes_info.append({
                'id': unique_id,
                'QR-codes-text': qr_code_text,
                'image_path': qr_image_path  # Consistent key
            })

        # Use the template and save the output
        excel_template_path = "Veriler/KaanEtiket.xlsx"  # Template file
        output_excel_path = "Veriler/qr_codes_output.xlsx"  # Output file
        create_excel_with_qr_codes(qr_codes_info, excel_template_path, output_excel_path)

        # Save the QR codes info to 'recent_qr_codes_file'
        df_recent_qr_codes = pd.DataFrame(qr_codes_info)
        df_recent_qr_codes.to_excel(recent_qr_codes_file, index=False)

        return qr_codes_info
    except Exception as e:
        st.error(f"Error generating QR codes: {e}")
        return []


def create_excel_with_qr_codes(qr_codes_info, excel_template_path, output_excel_path):
    """Insert QR code images and their IDs into a single Excel sheet using only odd rows."""
    try:
        # Load the template workbook
        wb = load_workbook(excel_template_path)
        ws = wb.active  # Use the active sheet or specify the sheet name if needed

        # Define odd rows for QR code placement
        max_rows = 999  # Maximum rows to consider; adjust based on your template's capacity
        qr_rows = [row for row in range(1, max_rows + 1, 2)]  # Odd rows only
        num_columns = 2  # Two columns (A and B)

        # Initialize QR code index
        qr_index = 0

        # Loop through QR rows and columns to place QR codes and IDs
        for row in qr_rows:
            for col in range(1, num_columns + 1):  # Columns A and B
                if qr_index >= len(qr_codes_info):
                    break  # Stop if all QR codes are placed

                qr_info = qr_codes_info[qr_index]
                id_ = qr_info['id']
                image_path = qr_info['image_path']  # Updated key name

                # Insert QR code image
                cell = ws.cell(row=row, column=col)
                img = XLImage(image_path)
                img.width, img.height = 100, 100  # Adjust size based on template's cell size
                img.anchor = cell.coordinate  # Place the image in the cell
                ws.add_image(img)

                # Write the ID below the image
                cell.value = id_
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

                qr_index += 1

            if qr_index >= len(qr_codes_info):
                break  # Exit once all QR codes are placed

        # Save the workbook
        wb.save(output_excel_path)
    except Exception as e:
        st.error(f"Error creating Excel with QR codes: {e}")


# ---------------------- Asset Management Functions ----------------------

# utils_1.py

def get_asset_history():
    """Fetch all asset history for display in the sidebar."""
    try:
        if os.path.exists(asset_movements_excel):
            movements_df = pd.read_excel(asset_movements_excel)

            # Sort by timestamp descending
            movements_df = movements_df.sort_values(by='Zaman', ascending=False)

            # Derive 'gönderen' and 'alıcı' based on 'Aksiyon' and existing columns
            movements_df['gönderen'] = movements_df.apply(
                lambda row: row['Çalışan'] if row['Aksiyon'] in ['Kullanıldı', 'İşlem İçin Gönderildi', 'removed'] else row[
                    'Firma'], axis=1
            )
            movements_df['alıcı'] = movements_df.apply(
                lambda row: row['Firma'] if row['Aksiyon'] == 'İşlem İçin Gönderildi' else row['Çalışan'] if row[
                                                                                                               'Aksiyon'] in [
                                                                                                               'Geri Alındı',
                                                                                                               'removed'] else 'N/A',
                axis=1
            )

            # Select relevant columns
            history = movements_df[['Varlık ID', 'varlık_adı', 'gönderen', 'alıcı', 'Zaman']].to_dict(orient='records')
            return history
        else:
            st.warning("Asset movements file not found.")
            return []
    except Exception as e:
        st.error(f"Error fetching asset history: {e}")
        return []


def delete_asset_from_excel(id_):
    """Delete an asset from the main Excel file based on its ID."""
    try:
        if os.path.exists(main_excel_file):
            assets_df = pd.read_excel(main_excel_file)
            if id_ in assets_df['id'].values:
                assets_df = assets_df[assets_df['id'] != id_]
                assets_df.to_excel(main_excel_file, index=False)
                st.success(f"Asset with ID {id_} has been deleted.")
                # Optionally, delete the corresponding QR code image
                qr_image_path = os.path.join(qr_codes_folder, f"{id_}.png")
                if os.path.exists(qr_image_path):
                    os.remove(qr_image_path)
            else:
                st.warning(f"No asset found with ID {id_}.")
        else:
            st.error("Main Excel file not found.")
    except Exception as e:
        st.error(f"Error deleting asset: {e}")


def get_current_stock_levels():
    """Fetch current stock levels for display."""
    try:
        if os.path.exists(main_excel_file):
            assets_df = pd.read_excel(main_excel_file)
            # Group by asset name and unit, sum the quantities
            stock_summary = assets_df.groupby(['varlık_adı', 'unit']).agg({'quantity': 'sum'}).reset_index()
            stock_summary.rename(columns={'quantity': 'Current Stock'}, inplace=True)
            return stock_summary
        else:
            st.warning("Main Excel file not found.")
            return pd.DataFrame()
    except Exception as e:
        st.error(f"Error fetching current stock levels: {e}")
        return pd.DataFrame()


def calculate_stock_summary():
    """Calculate a detailed stock summary, including total used."""
    try:
        if os.path.exists(main_excel_file) and os.path.exists(asset_movements_excel):
            assets_df = pd.read_excel(main_excel_file)
            movements_df = pd.read_excel(asset_movements_excel)

            # Calculate current stock
            current_stock = assets_df.groupby(['varlık_adı', 'unit']).agg({'quantity': 'sum'}).reset_index()
            current_stock.rename(columns={'quantity': 'Current Stock'}, inplace=True)

            # Calculate total used
            used_movements = movements_df[movements_df['Aksiyon'] == 'Kullanıldı']
            total_used = used_movements.groupby(['Varlık ID']).agg({'Miktar': 'sum'}).reset_index()
            total_used = total_used.merge(assets_df[['id', 'varlık_adı', 'unit']], left_on='Varlık ID', right_on='id',
                                          how='left')
            total_used_summary = total_used.groupby(['varlık_adı', 'unit']).agg({'Miktar': 'sum'}).reset_index()
            total_used_summary.rename(columns={'Miktar': 'Total Used'}, inplace=True)

            # Merge current stock and total used
            stock_summary = pd.merge(current_stock, total_used_summary, on=['varlık_adı', 'unit'], how='left')
            stock_summary['Total Used'] = stock_summary['Total Used'].fillna(0).astype(int)

            # Optionally, include minimum stock levels if available
            if os.path.exists(malzeme_uyari_file):
                malzeme_df = pd.read_excel(malzeme_uyari_file)
                stock_summary = pd.merge(stock_summary, malzeme_df[['varlık_adı', 'minimum_stock']], on='varlık_adı',
                                         how='left')
            else:
                stock_summary['minimum_stock'] = 0

            return stock_summary
        else:
            st.warning("Required Excel files not found.")
            return pd.DataFrame()
    except Exception as e:
        st.error(f"Error calculating stock summary: {e}")
        return pd.DataFrame()


def update_malzeme_uyari():
    """Update the Malzeme Uyarı.xlsx file based on current stock levels."""
    try:
        stock_summary = calculate_stock_summary()
        if not stock_summary.empty:
            # Filter assets below minimum stock
            low_stock = stock_summary[stock_summary['Current Stock'] <= stock_summary['minimum_stock']]
            if not low_stock.empty:
                # Write to Malzeme Uyarı.xlsx
                low_stock.to_excel(malzeme_uyari_file, index=False)
            else:
                # If no low stock, create an empty file with headers
                create_excel_file_if_missing(malzeme_uyari_file,
                                             ["varlık_adı", "unit", "Current Stock", "minimum_stock"])
        else:
            st.warning("Stock summary is empty. Cannot update Malzeme Uyarı.xlsx.")
    except Exception as e:
        st.error(f"Error updating Malzeme Uyarı.xlsx: {e}")


# ---------------------- QR Code Parsing and Asset Retrieval ----------------------

def parse_qr_code_data(qr_code_text):
    """Parse the QR code text to extract asset information."""
    try:
        # Split the qr_code_text by dashes
        parts = qr_code_text.split('-', 7)

        # Ensure the main parts are present
        if len(parts) < 8:
            st.error("QR code data is incomplete.")
            return None

        assetname = parts[0]
        gönderen = parts[1]
        alıcı = parts[2]
        miktar_unit = parts[3]
        adet = int(parts[4])
        kacıncı = int(parts[5])
        zaman = parts[6]

        # Remaining part is UUID (if present)
        uuid_part = parts[7] if len(parts) > 7 else None

        # Process miktar and unit
        miktar = ''.join(filter(str.isdigit, miktar_unit))
        unit = ''.join(filter(str.isalpha, miktar_unit))

        return {
            "assetname": assetname,
            "gönderen": gönderen,
            "alıcı": alıcı,
            "miktar": int(miktar),
            "unit": unit,
            "adet": adet,
            "kacıncı": kacıncı,
            "zaman": zaman,
            "uuid": uuid_part
        }
    except Exception as e:
        st.error(f"Error parsing QR code data: {e}")
        return None


def get_asset_by_qr(qr_code_text):
    """Fetch asset information from the Excel file using the QR code text."""
    try:
        if os.path.exists(main_excel_file):
            assets_df = pd.read_excel(main_excel_file)
            asset = assets_df[assets_df['qr_kodu'] == qr_code_text]
            if not asset.empty:
                return asset.iloc[0].to_dict()
        return None
    except Exception as e:
        st.error(f"Error fetching asset by QR code: {e}")
        return None


def get_asset_by_id(asset_id):
    """Fetch asset information from the Excel file using the asset ID."""
    try:
        if os.path.exists(main_excel_file):
            assets_df = pd.read_excel(main_excel_file)
            asset = assets_df[assets_df['id'] == asset_id]
            if not asset.empty:
                return asset.iloc[0].to_dict()
        return None
    except Exception as e:
        st.error(f"Error fetching asset by ID: {e}")
        return None


def get_asset_by_name(varlik_adi):
    """Fetch asset information from the Excel file using the asset name."""
    try:
        if os.path.exists(main_excel_file):
            assets_df = pd.read_excel(main_excel_file)
            asset = assets_df[assets_df['varlık_adı'] == varlik_adi]
            if not asset.empty:
                return asset.iloc[0].to_dict()
        return None
    except Exception as e:
        st.error(f"Error fetching asset by name: {e}")
        return None


# ---------------------- QR Code Generation ----------------------

def generate_qr_codes(assetname, unit, miktar, adet, gönderen, alıcı):
    """Generate QR codes and save them into the Excel template."""
    try:
        zaman = datetime.now().strftime("%Y-%m-%d-%H-%M-%S-%f")
        qr_codes_info = []

        for kacıncı in range(1, adet + 1):
            uuid_part = uuid.uuid4().hex
            unique_id = f"{assetname}-{gönderen}-{alıcı}-{miktar}{unit}-{adet}-{kacıncı}-{zaman}-{uuid_part}"
            qr_code_text = unique_id

            # Generate and save the QR code image
            qr = qrcode.make(qr_code_text)
            qr_image_path = os.path.join(qr_codes_folder, f"{unique_id}.png")
            qr.save(qr_image_path)

            # Collect QR code info with consistent key names
            qr_codes_info.append({
                'id': unique_id,
                'QR-codes-text': qr_code_text,
                'image_path': qr_image_path  # Consistent key
            })

        # Use the template and save the output
        excel_template_path = "Veriler/KaanEtiket.xlsx"  # Template file
        output_excel_path = "Veriler/qr_codes_output.xlsx"  # Output file
        create_excel_with_qr_codes(qr_codes_info, excel_template_path, output_excel_path)

        # Save the QR codes info to 'recent_qr_codes_file'
        df_recent_qr_codes = pd.DataFrame(qr_codes_info)
        df_recent_qr_codes.to_excel(recent_qr_codes_file, index=False)

        return qr_codes_info
    except Exception as e:
        st.error(f"Error generating QR codes: {e}")
        return []


# ---------------------- Adding Assets from Recent QR Codes ----------------------

def add_assets_from_recent_qr_codes():
    """Add assets from recent QR codes Excel file to main Excel."""
    try:
        if os.path.exists(recent_qr_codes_file):
            df_recent = pd.read_excel(recent_qr_codes_file)
            for index, row in df_recent.iterrows():
                id_ = row.get("id")
                qr_code_text = row.get("QR-codes-text")
                if not id_ or not qr_code_text:
                    st.warning(f"Missing data for row {index + 2}. Skipping.")
                    continue
                parts = qr_code_text.split('-', 7)
                if len(parts) < 8:
                    st.warning(f"QR code data is incomplete for ID: {id_}. Skipping.")
                    continue
                assetname, gönderen, alıcı, miktar_unit, adet, kacıncı, zaman, uuid_part = parts
                miktar = ''.join(filter(str.isdigit, miktar_unit))
                unit = ''.join(filter(str.isalpha, miktar_unit))
                try:
                    miktar = int(miktar)
                    adet = int(adet)
                    kacıncı = int(kacıncı)
                except ValueError:
                    st.warning(f"Miktar, adet veya kacıncı sayısal değil for ID: {id_}. Skipping.")
                    continue
                add_asset_to_excel(id_, qr_code_text, assetname, gönderen, alıcı, miktar, unit, adet, kacıncı, zaman)
            st.success("Tüm varlıklar başarıyla eklendi.")
        else:
            st.error("Recent QR codes file not found.")
    except Exception as e:
        st.error(f"Bir hata oluştu: {e}")


# ---------------------- Task Management Functions ----------------------

def get_task_history():
    """Fetch recent task history for display in the sidebar."""
    try:
        if os.path.exists(tasks_excel_file):
            tasks_df = pd.read_excel(tasks_excel_file)
            tasks_df = tasks_df.sort_values(by='created_at', ascending=False).head(10)
            tasks = tasks_df[['task_id', 'title', 'status', 'progress', 'created_at']].to_dict(orient='records')
            return tasks
        else:
            st.warning("Tasks Excel file not found.")
            return []
    except Exception as e:
        st.error(f"Error fetching task history: {e}")
        return []


def create_task(title, description, assigned_to, urgency):
    """Create a new task and save it to the Excel file."""
    try:
        if os.path.exists(tasks_excel_file):
            tasks_df = pd.read_excel(tasks_excel_file)
        else:
            tasks_df = pd.DataFrame(
                columns=["task_id", "title", "description", "assigned_to", "created_by", "urgency", "status",
                         "progress", "created_at", "updated_at"])
        task_id = tasks_df['task_id'].max() + 1 if not tasks_df.empty else 1
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        assigned_to_str = ",".join(assigned_to)
        new_task = {
            "task_id": task_id,
            "title": title,
            "description": description,
            "assigned_to": assigned_to_str,
            "created_by": st.session_state.username,
            "urgency": urgency,
            "status": "not seen",
            "progress": 0,
            "created_at": created_at,
            "updated_at": created_at
        }
        # Replace append() with pd.concat()
        tasks_df = pd.concat([tasks_df, pd.DataFrame([new_task])], ignore_index=True)
        tasks_df.to_excel(tasks_excel_file, index=False)
        return task_id
    except Exception as e:
        st.error(f"Error creating task: {e}")
        return None


def get_open_tasks():
    """Fetch open tasks for display, ordered by urgency and creation date."""
    try:
        if os.path.exists(tasks_excel_file):
            tasks_df = pd.read_excel(tasks_excel_file)
            tasks_df = tasks_df[tasks_df['status'] != 'completed']
            tasks_df = tasks_df.sort_values(by=['urgency', 'created_at'], ascending=[False, True])
            tasks = tasks_df.to_dict(orient='records')
            return tasks
        else:
            st.warning("Tasks Excel file not found.")
            return []
    except Exception as e:
        st.error(f"Error fetching open tasks: {e}")
        return []


def mark_task_as_seen(task_id, viewer):
    """Mark a task as seen, update its status, and notify the task creator."""
    try:
        if os.path.exists(tasks_excel_file):
            tasks_df = pd.read_excel(tasks_excel_file)
            task_index = tasks_df[tasks_df['task_id'] == task_id].index
            if not task_index.empty:
                index = task_index[0]
                tasks_df.at[index, 'status'] = 'in progress'
                tasks_df.at[index, 'updated_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                tasks_df.to_excel(tasks_excel_file, index=False)
                # Optionally, notify the task creator here
            else:
                st.warning(f"No task found with ID {task_id}.")
        else:
            st.error("Tasks Excel file not found.")
    except Exception as e:
        st.error(f"Error marking task as seen: {e}")


def get_task_creator(task_id):
    """Fetch the creator's username for a given task ID."""
    try:
        if os.path.exists(tasks_excel_file):
            tasks_df = pd.read_excel(tasks_excel_file)
            task = tasks_df[tasks_df['task_id'] == task_id]
            if not task.empty:
                return task.iloc[0]['created_by']
        return None
    except Exception as e:
        st.error(f"Error fetching task creator: {e}")
        return None


def update_task_progress(task_id, progress):
    """Update the progress of a task."""
    try:
        if os.path.exists(tasks_excel_file):
            tasks_df = pd.read_excel(tasks_excel_file)
            task_index = tasks_df[tasks_df['task_id'] == task_id].index
            if not task_index.empty:
                index = task_index[0]
                tasks_df.at[index, 'progress'] = progress
                tasks_df.at[index, 'updated_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                tasks_df.to_excel(tasks_excel_file, index=False)
            else:
                st.warning(f"No task found with ID {task_id}.")
        else:
            st.error("Tasks Excel file not found.")
    except Exception as e:
        st.error(f"Error updating task progress: {e}")


def complete_task(task_id):
    """Mark a task as completed."""
    try:
        if os.path.exists(tasks_excel_file):
            tasks_df = pd.read_excel(tasks_excel_file)
            task_index = tasks_df[tasks_df['task_id'] == task_id].index
            if not task_index.empty:
                index = task_index[0]
                tasks_df.at[index, 'status'] = 'completed'
                tasks_df.at[index, 'progress'] = 100
                tasks_df.at[index, 'updated_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                tasks_df.to_excel(tasks_excel_file, index=False)
            else:
                st.warning(f"No task found with ID {task_id}.")
        else:
            st.error("Tasks Excel file not found.")
    except Exception as e:
        st.error(f"Error completing task: {e}")


def get_completed_tasks():
    """Fetch completed tasks for display, ordered by completion date."""
    try:
        if os.path.exists(tasks_excel_file):
            tasks_df = pd.read_excel(tasks_excel_file)
            tasks_df = tasks_df[tasks_df['status'] == 'completed']
            tasks_df = tasks_df.sort_values(by='updated_at', ascending=False)
            tasks = tasks_df.to_dict(orient='records')
            return tasks
        else:
            st.warning("Tasks Excel file not found.")
            return []
    except Exception as e:
        st.error(f"Error fetching completed tasks: {e}")
        return []


# ---------------------- Asset Movement Functions ----------------------

def log_asset_movement(asset_id, action, quantity, partner_firm, worker, notes):
    """Log an asset movement."""
    try:
        if os.path.exists(asset_movements_excel):
            movements_df = pd.read_excel(asset_movements_excel)
        else:
            movements_df = pd.DataFrame(
                columns=["movement_id", "Zaman", "Varlık ID", "Varlık Adı", "Aksiyon", "Miktar", "Firma", "Çalışan",
                         "Notlar"])

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Get the asset name from main_excel_file
        assets_df = pd.read_excel(main_excel_file)
        asset_row = assets_df[assets_df['id'] == asset_id]
        varlik_adi = asset_row.iloc[0]['varlık_adı'] if not asset_row.empty else ''

        # Determine new movement_id
        movement_id = movements_df['movement_id'].max() + 1 if not movements_df.empty else 1

        # Append the new movement
        new_movement = {
            'movement_id': movement_id,
            'Zaman': timestamp,
            'Varlık ID': asset_id,
            'Varlık Adı': varlik_adi,
            'Aksiyon': action,
            'Miktar': quantity,
            'Firma': partner_firm,
            'Çalışan': worker,
            'Notlar': notes
        }
        # Replace append() with pd.concat()
        movements_df = pd.concat([movements_df, pd.DataFrame([new_movement])], ignore_index=True)
        movements_df.to_excel(asset_movements_excel, index=False)

        # Update the main Excel file
        update_asset_quantity_in_main_excel(asset_id, quantity if action == 'Geri Alındı' else -quantity)

        # Update Malzeme Uyarı.xlsx
        update_malzeme_uyari()
    except Exception as e:
        st.error(f"Error logging asset movement: {e}")


def undo_asset_movement(movement_id):
    """Undo a logged asset movement."""
    try:
        if os.path.exists(asset_movements_excel):
            movements_df = pd.read_excel(asset_movements_excel)
            # Find the movement details
            movement = movements_df[movements_df['movement_id'] == movement_id]
            if not movement.empty:
                movement = movement.iloc[0]
                asset_id = movement['Varlık ID']
                action = movement['Aksiyon']
                quantity = movement['Miktar']
                # Reverse the quantity change
                if action in ['Kullanıldı', 'İşlem İçin Gönderildi']:
                    update_asset_quantity_in_main_excel(asset_id, quantity)
                elif action == 'Geri Alındı':
                    update_asset_quantity_in_main_excel(asset_id, -quantity)
                # Delete the movement record
                movements_df = movements_df[movements_df['movement_id'] != movement_id]
                movements_df.to_excel(asset_movements_excel, index=False)
                # Update Malzeme Uyarı.xlsx after changes
                update_malzeme_uyari()
                st.success(f"Movement ID {movement_id} has been undone.")
                return True
            else:
                st.warning(f"No movement found with ID {movement_id}.")
        else:
            st.error("Asset movements Excel file not found.")
        return False
    except Exception as e:
        st.error(f"Error undoing asset movement: {e}")
        return False


def update_asset_quantity_in_main_excel(asset_id, quantity_change):
    """Update the quantity of an asset in the main Excel file."""
    try:
        if os.path.exists(main_excel_file):
            assets_df = pd.read_excel(main_excel_file)
            if asset_id in assets_df['id'].values:
                assets_df.loc[assets_df['id'] == asset_id, 'quantity'] += quantity_change
                # Ensure quantity doesn't go negative
                assets_df.loc[assets_df['quantity'] < 0, 'quantity'] = 0
                assets_df.to_excel(main_excel_file, index=False)
                # Update Malzeme Uyarı.xlsx after changes
                update_malzeme_uyari()
            else:
                st.warning(f"No asset found with ID {asset_id}.")
        else:
            st.error("Main Excel file not found.")
    except Exception as e:
        st.error(f"Error updating asset quantity: {e}")


# ---------------------- Task Management Functions Continued ----------------------

def get_personal_tasks(username):
    """Fetch tasks assigned to a specific user."""
    try:
        if os.path.exists(tasks_excel_file):
            tasks_df = pd.read_excel(tasks_excel_file)
            personal_tasks = tasks_df[tasks_df['assigned_to'].str.contains(username)]
            return personal_tasks
        else:
            st.warning("Tasks Excel file not found.")
            return pd.DataFrame()
    except Exception as e:
        st.error(f"Error fetching personal tasks: {e}")
        return pd.DataFrame()


# ---------------------- Notification Function ----------------------

def send_pushbullet_notification(title, message):
    """Send a Pushbullet notification."""
    try:
        import requests
        # Read Pushbullet API key from a config file or environment variable
        pushbullet_api_key = os.getenv("PUSHBULLET_API_KEY")
        if not pushbullet_api_key:
            st.warning("Pushbullet API key not found. Notifications will not be sent.")
            return
        data_send = {
            "type": "note",
            "title": title,
            "body": message
        }
        resp = requests.post(
            "https://api.pushbullet.com/v2/pushes",
            json=data_send,
            headers={"Access-Token": pushbullet_api_key}
        )
        if resp.status_code != 200:
            st.warning(f"Failed to send Pushbullet notification: {resp.text}")
    except Exception as e:
        st.error(f"Error sending Pushbullet notification: {e}")


# ---------------------- Export and Download Functions ----------------------

def convert_df_to_excel(df):
    """Convert a DataFrame to Excel bytes."""
    try:
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
        processed_data = output.getvalue()
        return processed_data
    except Exception as e:
        st.error(f"Error converting DataFrame to Excel: {e}")
        return None


# ---------------------- Cleanup and Maintenance Functions ----------------------

def reset_all_data():
    """Reset all data by clearing all Excel files."""
    try:
        # Reset main assets file
        create_excel_file_if_missing(main_excel_file,
                                     ["id", "qr_kodu", "gönderen", "alıcı", "varlık_adı", "miktar", "unit", "adet",
                                      "kacıncı", "zaman", "quantity"])

        # Reset recent QR codes file
        create_excel_file_if_missing(recent_qr_codes_file, ["id", "QR-codes-text", "image_path"])

        # Reset asset movements file
        create_excel_file_if_missing(asset_movements_excel,
                                     ["movement_id", "Zaman", "Varlık ID", "Varlık Adı", "Aksiyon", "Miktar", "Firma",
                                      "Çalışan", "Notlar"])

        # Reset tasks file
        create_excel_file_if_missing(tasks_excel_file,
                                     ["task_id", "title", "description", "assigned_to", "created_by", "urgency",
                                      "status", "progress", "created_at", "updated_at"])

        # Clear QR codes folder
        for filename in os.listdir(qr_codes_folder):
            file_path = os.path.join(qr_codes_folder, filename)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except Exception as e:
                st.error(f"Error deleting file {file_path}: {e}")

        st.success("All data has been reset successfully.")
    except Exception as e:
        st.error(f"Error resetting all data: {e}")


def clear_recent_qr_codes():
    """Clear the recent QR codes Excel file."""
    try:
        create_excel_file_if_missing(recent_qr_codes_file, ["id", "QR-codes-text", "image_path"])
        st.success("Recent QR codes have been cleared.")
    except Exception as e:
        st.error(f"Error clearing recent QR codes: {e}")


def get_filtered_movements(action_type, start_date, end_date):
    """Fetch movements based on filter criteria."""
    try:
        if os.path.exists(asset_movements_excel):
            movements_df = pd.read_excel(asset_movements_excel)
            #movements_df = get_asset_movements()
            # Filter by action type
            if action_type != "Hepsi":
                if action_type == "Kullanıldı":
                    movements_df = movements_df[movements_df['Aksiyon'] == 'Kullanıldı']
                elif action_type == "İşlem İçin Gönderildi":
                    movements_df = movements_df[movements_df['Aksiyon'] == 'İşlem İçin Gönderildi']
                elif action_type == "Geri Alındı":
                    movements_df = movements_df[movements_df['Aksiyon'] == 'Geri Alındı']

            # Filter by date range
            movements_df['Zaman'] = pd.to_datetime(movements_df['Zaman'])

            movements_df = movements_df[(movements_df['Zaman'] >= pd.to_datetime(start_date)) & (
                        movements_df['Zaman'] <= pd.to_datetime(end_date))]
            return movements_df
        else:
            st.warning("Asset movements Excel file not found.")
            #return pd.DataFrame()
    except Exception as e:
        st.error(f"Error fetching filtered movements: {e}")
        #return pd.DataFrame()

def get_asset_movements():
    """Return a DataFrame with asset movement logs."""
    if os.path.exists(asset_movements_excel):
        df = pd.read_excel(asset_movements_excel)
        return df
    else:
        return pd.DataFrame()


def get_filtered_assets(filter_option, filter_query):
    """Return a DataFrame of assets based on filter criteria."""
    if os.path.exists(main_excel_file):
        df = pd.read_excel(main_excel_file)
        if filter_option != "Hepsi" and filter_query:
            df = df[df[filter_option].str.contains(filter_query, case=False, na=False)]
        return df[['id', 'varlık_adı', 'gönderen', 'alıcı', 'miktar', 'unit']]
    else:
        return pd.DataFrame()


# ---------------------- Asset Removal and Undo Functions ----------------------

def remove_asset(asset_id):
    """Remove an asset from the main Excel file based on its ID."""
    try:
        if os.path.exists(main_excel_file):
            assets_df = pd.read_excel(main_excel_file)
            asset = assets_df[assets_df['id'] == asset_id]
            if not asset.empty:
                # Store the asset in session state for potential undo
                if 'last_deleted_asset' not in st.session_state:
                    st.session_state.last_deleted_asset = []
                st.session_state.last_deleted_asset.append(asset.to_dict(orient='records')[0])

                # Remove the asset
                assets_df = assets_df[assets_df['id'] != asset_id]
                assets_df.to_excel(main_excel_file, index=False)

                # Optionally, remove related QR code images
                qr_image_path = os.path.join(qr_codes_folder, f"{asset_id}.png")
                if os.path.exists(qr_image_path):
                    os.remove(qr_image_path)

                # Log the removal in asset movements
                log_asset_removal(asset_id, asset.iloc[0]['varlık_adı'])

                return True, "Asset successfully removed."
            else:
                return False, "Asset not found."
        else:
            return False, "Main Excel file not found."
    except Exception as e:
        return False, f"Error removing asset: {e}"


def log_asset_removal(asset_id, varlik_adi):
    """Log the asset removal in the asset movements Excel file."""
    try:
        if os.path.exists(asset_movements_excel):
            movements_df = pd.read_excel(asset_movements_excel)
        else:
            movements_df = pd.DataFrame(
                columns=["movement_id", "Zaman", "Varlık ID", "varlık_adı", "Aksiyon", "Miktar", "Firma", "Çalışan",
                         "Notlar"])

        movement_id = movements_df['movement_id'].max() + 1 if not movements_df.empty else 1
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        new_movement = {
            'movement_id': movement_id,
            'Zaman': timestamp,
            'Varlık ID': asset_id,
            'varlık_adı': varlik_adi,
            'Aksiyon': 'removed',
            'Miktar': 0,
            'Firma': 'N/A',
            'Çalışan': st.session_state.username,  # Assuming username is stored in session state
            'Notlar': 'Asset removed by admin.'
        }

        movements_df = pd.concat([movements_df, pd.DataFrame([new_movement])], ignore_index=True)
        movements_df.to_excel(asset_movements_excel, index=False)
    except Exception as e:
        st.error(f"Error logging asset removal: {e}")


def undo_last_deletion():
    """Undo the last asset deletion."""
    try:
        if 'last_deleted_asset' in st.session_state and st.session_state.last_deleted_asset:
            last_asset = st.session_state.last_deleted_asset.pop()
            add_asset_to_excel(
                id_=last_asset['id'],
                qr_kodu=last_asset['qr_kodu'],
                varlık_adı=last_asset['varlık_adı'],
                gönderen=last_asset['gönderen'],
                alıcı=last_asset['alıcı'],
                miktar=last_asset['miktar'],
                unit=last_asset['unit'],
                adet=last_asset['adet'],
                kacıncı=last_asset['kacıncı'],
                zaman=last_asset['zaman']
            )
            return True, "Last deletion has been undone."
        else:
            return False, "No deletions to undo."
    except Exception as e:
        return False, f"Error undoing deletion: {e}"


def get_all_assets():
    """Fetch all assets from the main Excel file without grouping."""
    try:
        if os.path.exists(main_excel_file):
            assets_df = pd.read_excel(main_excel_file)
            # Ensure 'id', 'varlık_adı', 'unit', 'quantity' columns exist
            required_columns = ['id', 'varlık_adı', 'unit', 'quantity']
            if all(col in assets_df.columns for col in required_columns):
                # Rename 'quantity' to 'Current Stock' for consistency
                assets_df = assets_df.rename(columns={'quantity': 'Current Stock'})
                return assets_df
            else:
                missing = [col for col in required_columns if col not in assets_df.columns]
                st.error(f"Missing columns in main Excel file: {missing}")
                return pd.DataFrame()
        else:
            st.warning("Main Excel file not found.")
            return pd.DataFrame()
    except Exception as e:
        st.error(f"Error fetching all assets: {e}")
        return pd.DataFrame()

def get_admin_usernames():
    """Fetch admin usernames from roles.txt."""
    try:
        with open("Veriler/roles.txt", "r") as file:
            user_data = json.load(file)  # Load data as JSON

        # Filter for users with the role 'admin'
        admins = [username for username, details in user_data['users'].items() if details['role'] == 'admin']
        return admins
    except json.JSONDecodeError:
        print("Error reading admin usernames from roles.txt: The file format is incorrect. Please ensure it is valid JSON.")
        return []
    except FileNotFoundError:
        print("Error: roles.txt file not found. Please ensure the file is in the correct directory.")
        return []